import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from config import OUTPUT_FILE


HEADERS = [
    "Report Number",
    "Report Type",
    "Date of Incident",
    "Accident Location",
    "Last Names Involved",
    "Jurisdiction",
    "Status",           # FOUND or NOT FOUND
    "Checked At",
]

# Color fills
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")   # found
RED_FILL   = PatternFill("solid", fgColor="FFC7CE")   # not found
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _get_or_create_workbook():
    """Load existing workbook or create a fresh one with headers."""
    if os.path.exists(OUTPUT_FILE):
        wb = openpyxl.load_workbook(OUTPUT_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Crash Reports"
        # Write headers
        for col, header in enumerate(HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
        # Column widths
        widths = [15, 20, 18, 30, 25, 25, 12, 20]
        for col, width in enumerate(widths, start=1):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = width
    return wb, ws


def save_found_report(record: dict):
    """Save a found report record to Excel."""
    wb, ws = _get_or_create_workbook()

    # Build accident location string
    street = record.get("street", "")
    cross  = record.get("crossStreet", "")
    location = f"{street} / {cross}".strip(" /") if cross else street

    # Last names as comma-separated string
    last_names = ", ".join(record.get("lastNames", []))

    row_data = [
        record.get("reportNumber", ""),
        record.get("reportTypeLabel", ""),
        record.get("dateOfIncident", ""),
        location,
        last_names,
        record.get("jurisdiction", ""),
        "FOUND",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ]

    row_num = ws.max_row + 1
    for col, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.fill = GREEN_FILL
        cell.alignment = Alignment(horizontal="center")

    wb.save(OUTPUT_FILE)
    print(f"  ✅ Saved FOUND report {record.get('reportNumber')} to Excel")


def save_not_found_report(report_number: str):
    """Save a not-found report number to Excel."""
    wb, ws = _get_or_create_workbook()

    row_data = [
        report_number,
        "", "", "", "", "",
        "NOT FOUND",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ]

    row_num = ws.max_row + 1
    for col, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_num, column=col, value=value)
        cell.fill = RED_FILL
        cell.alignment = Alignment(horizontal="center")

    wb.save(OUTPUT_FILE)
    print(f"  ❌ Saved NOT FOUND report {report_number} to Excel")


def get_summary():
    """Print a quick summary of results so far."""
    if not os.path.exists(OUTPUT_FILE):
        print("No results file yet.")
        return
    wb = openpyxl.load_workbook(OUTPUT_FILE)
    ws = wb.active
    found = 0
    not_found = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[6] == "FOUND":
            found += 1
        elif row[6] == "NOT FOUND":
            not_found += 1
    print(f"\n📊 Summary: {found} FOUND | {not_found} NOT FOUND | Total checked: {found + not_found}")